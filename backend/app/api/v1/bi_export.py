# backend/app/api/v1/bi_export.py
"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  LumindAd Enterprise · api/v1/bi_export.py
  Business Intelligence export endpoints  ← NUEVO

  Endpoints
  ──────────
  POST /api/v1/bi-export/powerbi    → xlsx  (StreamingResponse)
  POST /api/v1/bi-export/tableau    → csv   (StreamingResponse)
  POST /api/v1/bi-export/excel      → xlsx multi-sheet (StreamingResponse)
  POST /api/v1/bi-export/pdf        → pdf   (StreamingResponse)
  GET  /api/v1/bi-export/formats    → supported export formats

  BIExport module (frontend)
  ───────────────────────────
  Slide-in drawer from Analytics page (ExportReport.tsx).
  4 tools: Power BI · Tableau · Excel · PDF
  Source components: BIExport/PowerBIExport.tsx, TableauExport.tsx,
                     ExcelExport.tsx, PDFExport.tsx

  Data exported (seed from LumindAd.jsx)
  ─────────────────────────────────────────
  campaigns:  6 rows  (lines 103–110)
  analytics:  7 rows  (lines 112–118)
  budget:     7 days  (lines 119–123)
  platforms:  5 rows  (lines 95–101)
  ml_models:  4 rows  (lines 630–635)

  Power BI format
  ────────────────
  Multi-sheet Excel workbook (.xlsx) + metadata JSON header.
  Sheet layout:
    BI_Metadata  — workspace_id, dataset_name, refresh_mode, generated_by
    Campaigns    — 6 campaigns with all fields
    Analytics    — 7-point time-series
    Budget       — 7-day daily spend
    Platforms    — 5 platform shares
    ML_Models    — 4 model cards

  Tableau format
  ───────────────
  CSV file (Tableau Web Data Connector compatible).
  All data tables concatenated with a 'table_name' discriminator column.

  Excel format
  ─────────────
  Multi-sheet .xlsx: one sheet per domain.

  PDF format
  ───────────
  Plain-text PDF report using reportlab (graceful fallback to text if
  reportlab is not installed).

  Author : Elizabeth Díaz Familia
           AI Data Scientist · Sustainable Intelligence & BI
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Literal, Optional

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from app.config import settings
from app.dependencies import AuthUser, get_current_user

logger = logging.getLogger(__name__)
router = APIRouter()

# ─── Optional heavy deps ──────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False
    Workbook = None  # type: ignore[assignment, misc]

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas as rl_canvas
    _REPORTLAB = True
except ImportError:
    _REPORTLAB = False


# ═══════════════════════════════════════════════════════════════
# SEED DATA — exact values from LumindAd.jsx / BIExport/data.ts
# ═══════════════════════════════════════════════════════════════

# campaigns (LumindAd.jsx lines 103–110)
_CAMPAIGNS = [
    {"id": "C-001", "name": "Summer Sale 2025",    "platform": "Google Ads", "status": "active",    "budget": 5000, "spent": 3240, "impressions": 124500, "clicks": 8920,  "ctr": "7.16%", "conv": 342, "roas": 3.8},
    {"id": "C-002", "name": "Brand Awareness Q1",  "platform": "Meta Ads",   "status": "active",    "budget": 8000, "spent": 5180, "impressions": 287000, "clicks": 12400, "ctr": "4.32%", "conv": 520, "roas": 2.9},
    {"id": "C-003", "name": "Product Launch Beta", "platform": "TikTok",     "status": "paused",    "budget": 3500, "spent": 1890, "impressions": 98200,  "clicks": 5430,  "ctr": "5.53%", "conv": 187, "roas": 4.2},
    {"id": "C-004", "name": "Retargeting Dec",     "platform": "Google Ads", "status": "active",    "budget": 2000, "spent": 1740, "impressions": 43100,  "clicks": 3280,  "ctr": "7.61%", "conv": 245, "roas": 5.1},
    {"id": "C-005", "name": "LinkedIn B2B Push",   "platform": "LinkedIn",   "status": "draft",     "budget": 6000, "spent": 0,    "impressions": 0,      "clicks": 0,     "ctr": "—",     "conv": 0,   "roas": 0.0},
    {"id": "C-006", "name": "Holiday Promos",      "platform": "Meta Ads",   "status": "completed", "budget": 4200, "spent": 4198, "impressions": 178000, "clicks": 9870,  "ctr": "5.54%", "conv": 430, "roas": 3.5},
]

# analytics time-series (LumindAd.jsx lines 112–118)
_ANALYTICS = [
    {"date": "Jan 1",  "impressions": 11000, "clicks": 780,  "conversions": 38},
    {"date": "Jan 8",  "impressions": 15200, "clicks": 1120, "conversions": 67},
    {"date": "Jan 15", "impressions": 18700, "clicks": 1480, "conversions": 89},
    {"date": "Jan 22", "impressions": 22100, "clicks": 1830, "conversions": 118},
    {"date": "Jan 29", "impressions": 24800, "clicks": 2150, "conversions": 142},
    {"date": "Feb 5",  "impressions": 27300, "clicks": 2480, "conversions": 168},
    {"date": "Feb 12", "impressions": 30100, "clicks": 2820, "conversions": 198},
]

# budget daily (LumindAd.jsx lines 119–123)
_BUDGET = [
    {"day": "Mon", "budget": 1500, "spend": 1240},
    {"day": "Tue", "budget": 1500, "spend": 1820},
    {"day": "Wed", "budget": 1500, "spend": 1470},
    {"day": "Thu", "budget": 1500, "spend": 2250},
    {"day": "Fri", "budget": 1500, "spend": 2480},
    {"day": "Sat", "budget": 1500, "spend": 1840},
    {"day": "Sun", "budget": 1500, "spend": 1350},
]

# platforms (LumindAd.jsx lines 95–101)
_PLATFORMS = [
    {"name": "Google Ads", "value": 38, "color": "#4285f4"},
    {"name": "Meta Ads",   "value": 29, "color": "#1877f2"},
    {"name": "TikTok",     "value": 18, "color": "#ff0050"},
    {"name": "LinkedIn",   "value": 10, "color": "#0077b5"},
    {"name": "Twitter/X",  "value": 5,  "color": "#1da1f2"},
]

# ML models (LumindAd.jsx lines 630–635)
_ML_MODELS = [
    {"name": "Churn Predictor",  "algorithm": "XGBoost",          "accuracy": 87.3, "status": "active"},
    {"name": "Anomaly Detector", "algorithm": "Isolation Forest",  "accuracy": 94.1, "status": "active"},
    {"name": "Click Predictor",  "algorithm": "Neural Network",    "accuracy": 82.7, "status": "active"},
    {"name": "ROAS Optimizer",   "algorithm": "AutoML",            "accuracy": 91.2, "status": "training"},
]


def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _generated_by(user: AuthUser) -> str:
    return f"{user.name} ({user.email})"


# ═══════════════════════════════════════════════════════════════
# PYDANTIC SCHEMAS
# ═══════════════════════════════════════════════════════════════

class PowerBIExportRequest(BaseModel):
    datasetName:        str   = Field(default="LumindAd_Dataset", description="Power BI dataset name")
    workspaceId:        Optional[str] = None
    refreshMode:        Literal["full", "incremental"] = "full"
    includePredictions: bool  = True
    dateRange:          Optional[str] = None


class TableauExportRequest(BaseModel):
    workbookName:       str   = Field(default="LumindAd_Workbook")
    dataSourceType:     Literal["extract", "live"] = "extract"
    includePredictions: bool  = True


class ExcelExportRequest(BaseModel):
    reportName:         str   = Field(default="LumindAd_Report")
    includeCharts:      bool  = True
    includePredictions: bool  = True


class PDFExportRequest(BaseModel):
    title:      str = Field(default="LumindAd Enterprise Report")
    subtitle:   Optional[str] = "Analytics & Performance Dashboard"
    includeML:  bool = True


class ExportFormat(BaseModel):
    id:          str
    name:        str
    description: str
    extension:   str
    contentType: str
    icon:        str


# ═══════════════════════════════════════════════════════════════
# XLSX BUILDER HELPERS
# ═══════════════════════════════════════════════════════════════

_HEADER_FILL   = "7C3AED"   # violet-600 — primary brand colour
_HEADER_FONT   = "FFFFFF"   # white text on header

def _apply_header_style(ws, row: int, cols: int) -> None:
    """Apply violet header style to first row."""
    if not _OPENPYXL:
        return
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    font = Font(bold=True, color=_HEADER_FONT)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _write_table(ws, headers: List[str], rows: List[dict], start_row: int = 1) -> None:
    """Write headers + data rows to a worksheet."""
    # Header row
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    _apply_header_style(ws, start_row, len(headers))

    # Data rows
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, key in enumerate(headers, 1):
            val = row.get(key, row.get(key.lower(), ""))
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Auto-width (approximate)
    if _OPENPYXL:
        for col_cells in ws.iter_cols(min_row=start_row, max_row=start_row + len(rows)):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)


def _build_xlsx_workbook(report_name: str, user: AuthUser) -> io.BytesIO:
    """Build a multi-sheet Excel workbook with all LumindAd data."""
    if not _OPENPYXL:
        # Fallback: return a minimal CSV-in-xlsx placeholder
        buf = io.BytesIO()
        buf.write(b"openpyxl not installed")
        buf.seek(0)
        return buf

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Metadata sheet ────────────────────────────────────────
    meta = wb.create_sheet("Metadata")
    meta_rows = [
        ["Property",       "Value"],
        ["Report Name",    report_name],
        ["Generated By",   _generated_by(user)],
        ["Generated At",   _now_str()],
        ["App Version",    settings.APP_VERSION],
        ["Total Campaigns", len(_CAMPAIGNS)],
        ["Total Analytics Points", len(_ANALYTICS)],
        ["Green AI Scope", "GHG Scope 2"],
        ["Carbon Intensity", f"{settings.CARBON_INTENSITY_KG_KWH} kgCO₂/kWh"],
        ["TelecomX Compatible", "Yes"],
    ]
    for r_idx, row in enumerate(meta_rows, 1):
        for c_idx, val in enumerate(row, 1):
            meta.cell(row=r_idx, column=c_idx, value=val)
    _apply_header_style(meta, 1, 2)
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 42

    # ── Campaigns sheet ───────────────────────────────────────
    ws_c = wb.create_sheet("Campaigns")
    _write_table(ws_c, ["id", "name", "platform", "status", "budget", "spent", "impressions", "clicks", "ctr", "conv", "roas"], _CAMPAIGNS)

    # ── Analytics sheet ───────────────────────────────────────
    ws_a = wb.create_sheet("Analytics")
    _write_table(ws_a, ["date", "impressions", "clicks", "conversions"], _ANALYTICS)

    # ── Budget sheet ──────────────────────────────────────────
    ws_b = wb.create_sheet("Budget")
    _write_table(ws_b, ["day", "budget", "spend"], _BUDGET)

    # ── Platforms sheet ───────────────────────────────────────
    ws_p = wb.create_sheet("Platforms")
    _write_table(ws_p, ["name", "value", "color"], _PLATFORMS)

    # ── ML Models sheet ───────────────────────────────────────
    ws_m = wb.create_sheet("ML_Models")
    _write_table(ws_m, ["name", "algorithm", "accuracy", "status"], _ML_MODELS)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@router.get(
    "/formats",
    response_model=List[ExportFormat],
    summary="List supported export formats",
)
async def list_formats(
    current_user: AuthUser = Depends(get_current_user),
) -> List[ExportFormat]:
    """
    Returns the 4 supported BI export formats with metadata.
    Used by BIExport/index.tsx to render the format selector.
    """
    return [
        ExportFormat(
            id="powerbi",
            name="Power BI",
            description="Multi-sheet Excel workbook with BI metadata, ready for Power BI Desktop import",
            extension=".xlsx",
            contentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            icon="📊",
        ),
        ExportFormat(
            id="tableau",
            name="Tableau",
            description="CSV export compatible with Tableau Web Data Connector",
            extension=".csv",
            contentType="text/csv",
            icon="📈",
        ),
        ExportFormat(
            id="excel",
            name="Excel",
            description="Multi-sheet Excel workbook: Campaigns, Analytics, Budget, ML Models",
            extension=".xlsx",
            contentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            icon="📋",
        ),
        ExportFormat(
            id="pdf",
            name="PDF Report",
            description="Formatted PDF report with KPIs, charts, and ML model summary",
            extension=".pdf",
            contentType="application/pdf",
            icon="📄",
        ),
    ]


@router.post(
    "/powerbi",
    summary="Export data for Power BI",
    response_description="Multi-sheet .xlsx workbook with BI_Metadata sheet",
)
async def export_powerbi(
    body:         PowerBIExportRequest,
    current_user: AuthUser = Depends(get_current_user),
) -> StreamingResponse:
    """
    Export all LumindAd data as a Power BI-compatible Excel workbook.

    Sheets:
      • Metadata     — workspace_id, dataset_name, refresh_mode
      • Campaigns    — 6 campaigns (seed data from LumindAd.jsx)
      • Analytics    — 7-point time-series
      • Budget       — 7-day daily spend vs budget
      • Platforms    — 5 platform shares
      • ML_Models    — 4 model cards

    The returned .xlsx can be imported directly into Power BI Desktop
    via Home → Get Data → Excel Workbook.

    Configured for dataset: `{body.datasetName}` / workspace: `{body.workspaceId or 'default'}`
    """
    logger.info(
        "Power BI export requested by %s: dataset=%s mode=%s",
        current_user.email, body.datasetName, body.refreshMode,
    )

    buf = _build_xlsx_workbook(body.datasetName, current_user)

    ts  = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"LumindAd_PowerBI_{body.datasetName.replace(' ', '_')}_{ts}.xlsx"

    return StreamingResponse(
        content   = iter([buf.read()]),
        media_type= "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers   = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Format":     "powerbi",
            "X-Dataset-Name":      body.datasetName,
            "X-Generated-By":      current_user.email,
        },
    )


@router.post(
    "/tableau",
    summary="Export data for Tableau",
    response_description="CSV file compatible with Tableau Web Data Connector",
)
async def export_tableau(
    body:         TableauExportRequest,
    current_user: AuthUser = Depends(get_current_user),
) -> StreamingResponse:
    """
    Export all LumindAd data as a Tableau-compatible CSV.

    Each row includes a `table_name` discriminator column so Tableau
    can split the data into separate logical tables via calculated fields.

    Tables included:
      campaigns | analytics | budget | platforms | ml_models

    Compatible with Tableau Web Data Connector (WDC) and
    Tableau Desktop File → Open (CSV).
    """
    logger.info(
        "Tableau export requested by %s: workbook=%s",
        current_user.email, body.workbookName,
    )

    # Build combined CSV with table_name discriminator
    output = io.StringIO()
    writer = csv.writer(output)

    # Header — superset of all table columns
    header = ["table_name", "id", "name", "platform", "status", "budget",
              "spent", "impressions", "clicks", "ctr", "conv", "roas",
              "date", "conversions", "day", "spend", "value", "color",
              "algorithm", "accuracy"]
    writer.writerow(header)

    def _pad(d: dict, keys: List[str]) -> List[Any]:
        return [d.get(k, "") for k in keys]

    for row in _CAMPAIGNS:
        writer.writerow(["campaigns"] + _pad(
            row, ["id", "name", "platform", "status", "budget", "spent",
                  "impressions", "clicks", "ctr", "conv", "roas",
                  "", "", "", "", "", "", "", ""]))

    for row in _ANALYTICS:
        writer.writerow(["analytics", "", "", "", "", "", "",
                         row.get("impressions", ""), row.get("clicks", ""), "", "",
                         "", row.get("date", ""), row.get("conversions", ""),
                         "", "", "", "", "", ""])

    for row in _BUDGET:
        writer.writerow(["budget"] + [""] * 8 + ["", "", "", "", "",
                         row.get("day", ""), row.get("spend", ""),
                         row.get("budget", ""), "", "", ""])

    for row in _PLATFORMS:
        writer.writerow(["platforms", "", row.get("name", ""), "", "", "",
                         "", "", "", "", "", "",
                         "", "", "", "", row.get("value", ""), row.get("color", ""), "", ""])

    for row in _ML_MODELS:
        writer.writerow(["ml_models", "", row.get("name", ""), "", "",
                         "", "", "", "", "", "", "",
                         "", "", "", "", "", "", row.get("algorithm", ""), row.get("accuracy", "")])

    csv_bytes = output.getvalue().encode("utf-8-sig")   # BOM for Excel compatibility

    ts       = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"LumindAd_Tableau_{body.workbookName.replace(' ', '_')}_{ts}.csv"

    return StreamingResponse(
        content    = iter([csv_bytes]),
        media_type = "text/csv",
        headers    = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Format":     "tableau",
            "X-Workbook-Name":     body.workbookName,
        },
    )


@router.post(
    "/excel",
    summary="Export multi-sheet Excel report",
)
async def export_excel(
    body:         ExcelExportRequest,
    current_user: AuthUser = Depends(get_current_user),
) -> StreamingResponse:
    """
    Export a clean multi-sheet Excel workbook.

    Sheets: Campaigns · Analytics · Budget · Platforms · ML_Models · Metadata

    Each sheet has:
      - Violet (#7C3AED) header row (brand colour)
      - Auto-sized column widths
      - All seed data from LumindAd.jsx

    The workbook is production-ready for sharing with stakeholders
    who do not have Power BI or Tableau licences.
    """
    logger.info("Excel export requested by %s: report=%s", current_user.email, body.reportName)

    buf      = _build_xlsx_workbook(body.reportName, current_user)
    ts       = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"LumindAd_{body.reportName.replace(' ', '_')}_{ts}.xlsx"

    return StreamingResponse(
        content    = iter([buf.read()]),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Format":     "excel",
            "X-Report-Name":       body.reportName,
        },
    )


@router.post(
    "/pdf",
    summary="Export PDF analytics report",
)
async def export_pdf(
    body:         PDFExportRequest,
    current_user: AuthUser = Depends(get_current_user),
) -> StreamingResponse:
    """
    Generate a formatted PDF analytics report.

    Sections:
      1. Title page — report name, author, date
      2. KPI Summary — Total Impressions, CTR, Conv Rate, CPC
      3. Campaigns table — 6 campaigns
      4. Budget Summary — $28,500 total / $18,347 spent / 64%
      5. ML Models — 4 model cards with accuracy
      6. Green AI — CO₂ badge, GHG Scope 2

    Uses reportlab if available; falls back to plain UTF-8 text PDF.
    """
    logger.info("PDF export requested by %s", current_user.email)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if _REPORTLAB:
        buf = _build_pdf_reportlab(body, current_user)
    else:
        buf = _build_pdf_text_fallback(body, current_user)

    filename = f"LumindAd_Report_{ts}.pdf"

    return StreamingResponse(
        content    = iter([buf.read()]),
        media_type = "application/pdf",
        headers    = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Format":     "pdf",
        },
    )


# ─── PDF builders ────────────────────────────────────────────────────────────

def _build_pdf_reportlab(body: PDFExportRequest, user: AuthUser) -> io.BytesIO:
    """reportlab PDF — structured report with sections."""
    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=letter)
    w, h = letter
    y    = h - inch

    # ── Title page ────────────────────────────────────────────
    c.setFont("Helvetica-Bold", 22)
    c.setFillColorRGB(0.486, 0.227, 0.929)   # #7C3AED violet-600
    c.drawCentredString(w / 2, y, body.title)
    y -= 28

    if body.subtitle:
        c.setFont("Helvetica", 13)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawCentredString(w / 2, y, body.subtitle)
        y -= 20

    c.setFont("Helvetica", 10)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawCentredString(w / 2, y, f"Generated by {user.name}  ·  {_now_str()}")
    y -= 14
    c.drawCentredString(w / 2, y, f"LumindAd Enterprise v{settings.APP_VERSION}  ·  GHG Scope 2 Certified")
    y -= 30

    # Separator line
    c.setStrokeColorRGB(0.486, 0.227, 0.929)
    c.setLineWidth(1.5)
    c.line(inch, y, w - inch, y)
    y -= 24

    def _section(title: str) -> None:
        nonlocal y
        if y < 150:
            c.showPage()
            nonlocal_y = h - inch
            y = nonlocal_y
        c.setFont("Helvetica-Bold", 13)
        c.setFillColorRGB(0.486, 0.227, 0.929)
        c.drawString(inch, y, title)
        y -= 18
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica", 10)

    def _line(text: str, indent: int = 0) -> None:
        nonlocal y
        if y < 80:
            c.showPage()
            y = h - inch
        c.drawString(inch + indent, y, text)
        y -= 14

    # ── KPI Summary ───────────────────────────────────────────
    _section("📊 KPI Summary")
    _line("Total Impressions:   531,200  (+24.5%)")
    _line("Click-Through Rate:    7.32%  (+12.3%)")
    _line("Conversion Rate:       4.18%  (+8.7%)")
    _line("Cost Per Click:        $1.24  (-5.2%)")
    y -= 8

    # ── Campaigns ─────────────────────────────────────────────
    _section("🎯 Campaigns (6 active)")
    for camp in _CAMPAIGNS:
        _line(f"  [{camp['status'].upper():<10}]  {camp['name']:<28} {camp['platform']:<12}  "
              f"Spent ${camp['spent']:>7,.0f}  ROAS {camp['roas']}")
    y -= 8

    # ── Budget ────────────────────────────────────────────────
    _section("💰 Budget Summary")
    _line("Total Budget:   $28,500")
    _line("Total Spent:    $18,347  (+18.2% vs prior period)")
    _line("Remaining:      $10,153")
    _line("Budget Utilised:   64%")
    _line("AI Recommendation: Reallocate $1,200 from Meta → Google Ads (+23% ROAS)")
    y -= 8

    # ── ML Models ─────────────────────────────────────────────
    if body.includeML:
        _section("🤖 ML Models")
        for m in _ML_MODELS:
            _line(f"  {m['name']:<22}  {m['algorithm']:<20}  Accuracy: {m['accuracy']}%  [{m['status']}]")
        y -= 8

    # ── Green AI ──────────────────────────────────────────────
    _section("🌱 Green AI — GHG Scope 2")
    _line(f"Carbon Intensity:  {settings.CARBON_INTENSITY_KG_KWH} kgCO₂/kWh (IEA 2023 global avg)")
    _line(f"Data Centre PUE:   {settings.PUE}")
    _line(f"CPU Power:         {settings.CPU_POWER_W} W  ·  GPU: {settings.GPU_POWER_W} W (NVIDIA T4)")
    _line("Session Badge:     0.003 gCO₂ · GHG Scope 2")

    c.save()
    buf.seek(0)
    return buf


def _build_pdf_text_fallback(body: PDFExportRequest, user: AuthUser) -> io.BytesIO:
    """
    Minimal valid PDF — plain text embedded in raw PDF syntax.
    Used when reportlab is not installed.
    """
    lines = [
        body.title,
        body.subtitle or "",
        f"Generated by {user.name}  ·  {_now_str()}",
        f"LumindAd Enterprise v{settings.APP_VERSION}",
        "",
        "=== KPI SUMMARY ===",
        "Total Impressions:  531,200  (+24.5%)",
        "Click-Through Rate:   7.32%  (+12.3%)",
        "Conversion Rate:      4.18%  (+8.7%)",
        "Cost Per Click:       $1.24  (-5.2%)",
        "",
        "=== CAMPAIGNS ===",
    ] + [
        f"[{c['status'].upper():<10}]  {c['name']:<28}  {c['platform']:<12}  "
        f"Spent ${c['spent']:>7,.0f}  ROAS {c['roas']}"
        for c in _CAMPAIGNS
    ] + [
        "",
        "=== BUDGET SUMMARY ===",
        "Total Budget:  $28,500",
        "Total Spent:   $18,347  (+18.2%)",
        "Remaining:     $10,153  (36%)",
        "",
        "=== ML MODELS ===",
    ] + [
        f"{m['name']:<22}  {m['algorithm']:<20}  Acc: {m['accuracy']}%  [{m['status']}]"
        for m in _ML_MODELS
    ] + [
        "",
        "=== GREEN AI — GHG SCOPE 2 ===",
        f"Carbon Intensity: {settings.CARBON_INTENSITY_KG_KWH} kgCO2/kWh",
        f"PUE: {settings.PUE}  CPU: {settings.CPU_POWER_W}W  GPU: {settings.GPU_POWER_W}W",
        "Session Badge: 0.003 gCO2 · GHG Scope 2",
    ]

    text_content = "\n".join(lines)

    # Minimal raw PDF envelope
    pdf = (
        "%PDF-1.4\n"
        "1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
        "2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
        "3 0 obj\n<< /Type /Page /Parent 2 0 R /Resources << /Font << /F1 4 0 R >> >>"
        " /MediaBox [0 0 612 792] /Contents 5 0 R >>\nendobj\n"
        "4 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>\nendobj\n"
    )
    stream_lines = []
    y = 750
    stream_lines.append("BT")
    stream_lines.append("/F1 9 Tf")
    for line in lines[:60]:   # truncate to one page
        escaped = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        stream_lines.append(f"72 {y} Td ({escaped}) Tj 0 -13 Td")
        y -= 13
    stream_lines.append("ET")
    stream_body = "\n".join(stream_lines).encode("latin-1", errors="replace")

    pdf += f"5 0 obj\n<< /Length {len(stream_body)} >>\nstream\n"
    pdf_bytes = pdf.encode("latin-1") + stream_body + b"\nendstream\nendobj\n"
    pdf_bytes += b"xref\n0 6\n0000000000 65535 f \n%%EOF\n"

    buf = io.BytesIO(pdf_bytes)
    buf.seek(0)
    return buf
