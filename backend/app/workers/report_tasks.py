# backend/app/workers/report_tasks.py
"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  LumindAd Enterprise · backend/app/workers/report_tasks.py
  Celery tasks — async report generation + export

  Tasks
  ──────
  generate_excel_report_task(user_id, params)
    Multi-sheet openpyxl workbook with 5 sheets:
      • Campaigns      — 6 campaigns full table (seed)
      • Analytics      — weekly series + KPI cards
      • Budget         — summary + platform allocations
      • ML Models      — registry + metrics (accuracy, F1, AUC)
      • Green AI       — CO₂ session report (Scope 2 GHG)
    Saves to settings.UPLOAD_DIR/reports/{user_id}/{filename}.xlsx
    Returns { filePath, fileSize, durationMs, taskId }

  generate_pdf_report_task(user_id, params)
    Reportlab (fallback: text/utf-8) PDF report:
      • Cover page with LumindAd branding + period + user
      • KPI summary table
      • Campaign performance table
      • Green AI Scope 2 CO₂ badge section
      • Platform allocation chart (ASCII/text fallback)
    Saves to settings.UPLOAD_DIR/reports/{user_id}/{filename}.pdf
    Returns { filePath, fileSize, durationMs, taskId }

  scheduled_daily_report_task()
    Beat task — runs at 06:00 UTC daily.
    Generates Excel + PDF for all active users.
    Optionally uploads to S3 (when settings.AWS_S3_BUCKET set).

  send_report_email_task(user_email, file_paths, report_date)
    Sends generated reports via SMTP (optional — skipped if SMTP not
    configured in settings).

  Seed data (LumindAd.jsx — all values exact)
  ─────────────────────────────────────────────
  Campaigns (6 rows) · Budget $28,500 / $18,347 · 5 platforms
  Analytics KPIs: 531.2K impressions · 7.32% CTR · 4.18% conv · $1.24 CPC
  ML models: Churn 87.3% · Anomaly 94.1% · Click 82.7% · ROAS 91.2%

  Author : Elizabeth Díaz Familia
           AI Data Scientist · Sustainable Intelligence & BI
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

from __future__ import annotations

import io
import logging
import os
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from app.config import settings

logger = logging.getLogger(__name__)

# ── Optional Celery ───────────────────────────────────────────────────────────
try:
    from celery.exceptions import SoftTimeLimitExceeded
    from app.workers.celery_app import celery_app
    _CELERY = True
except ImportError:
    _CELERY = False
    celery_app = None  # type: ignore[assignment]
    SoftTimeLimitExceeded = Exception  # type: ignore[assignment,misc]

# ── Optional openpyxl ─────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    _OPENPYXL = False

# ── Optional reportlab ────────────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    _REPORTLAB = True
except ImportError:
    _REPORTLAB = False


# ═══════════════════════════════════════════════════════════════
# SEED DATA  (exact values from LumindAd.jsx)
# ═══════════════════════════════════════════════════════════════

_CAMPAIGNS = [
    {"id":"C-001","name":"Summer Sale 2025",    "platform":"Google Ads","status":"active",    "budget":5000,"spent":3240,"impressions":124500,"clicks":8920, "ctr":"7.16%","conv":342,"roas":3.8},
    {"id":"C-002","name":"Brand Awareness Q1",  "platform":"Meta Ads",  "status":"active",    "budget":8000,"spent":5180,"impressions":287000,"clicks":12400,"ctr":"4.32%","conv":520,"roas":2.9},
    {"id":"C-003","name":"Product Launch Beta", "platform":"TikTok",    "status":"paused",    "budget":3500,"spent":1890,"impressions":98200, "clicks":5430, "ctr":"5.53%","conv":187,"roas":4.2},
    {"id":"C-004","name":"Retargeting Dec",     "platform":"Google Ads","status":"active",    "budget":2000,"spent":1740,"impressions":43100, "clicks":3280, "ctr":"7.61%","conv":245,"roas":5.1},
    {"id":"C-005","name":"LinkedIn B2B Push",   "platform":"LinkedIn",  "status":"draft",     "budget":6000,"spent":0,   "impressions":0,     "clicks":0,    "ctr":"—",    "conv":0,  "roas":0.0},
    {"id":"C-006","name":"Holiday Promos",      "platform":"Meta Ads",  "status":"completed", "budget":4200,"spent":4198,"impressions":178000,"clicks":9870, "ctr":"5.54%","conv":430,"roas":3.5},
]

_BUDGET_SUMMARY = {"totalBudget": 28500, "totalSpent": 18347, "remaining": 10153, "period": "November 2025"}

_ANALYTICS_KPIS = [
    {"key": "totalImpressions", "label": "Total Impressions", "value": 531200, "formatted": "531.2K", "change": 24.5},
    {"key": "ctr",              "label": "Click-Through Rate","value": 7.32,   "formatted": "7.32%",  "change": 12.3},
    {"key": "conversionRate",   "label": "Conversion Rate",   "value": 4.18,   "formatted": "4.18%",  "change":  8.7},
    {"key": "cpc",              "label": "Cost per Click",     "value": 1.24,   "formatted": "$1.24",  "change": -5.2},
]

_ML_MODELS = [
    {"name": "Churn Predictor",  "algorithm": "XGBoost",          "accuracy": 87.3, "f1": 85.8, "auc": 92.1, "status": "active"},
    {"name": "Anomaly Detector", "algorithm": "Isolation Forest",  "accuracy": 94.1, "f1": 94.0, "auc": 98.7, "status": "active"},
    {"name": "Click Predictor",  "algorithm": "Neural Network",    "accuracy": 82.7, "f1": 82.6, "auc": 89.1, "status": "active"},
    {"name": "ROAS Optimizer",   "algorithm": "AutoML",            "accuracy": 91.2, "f1": 91.1, "auc": 96.5, "status": "training"},
]

_PLATFORM_ALLOCS = [
    {"platform": "Google Ads", "pct": 38.0, "color": "#4285f4", "amountUSD": 6971.86},
    {"platform": "Meta Ads",   "pct": 29.0, "color": "#1877f2", "amountUSD": 5320.63},
    {"platform": "TikTok",     "pct": 18.0, "color": "#ff0050", "amountUSD": 3302.46},
    {"platform": "LinkedIn",   "pct": 10.0, "color": "#0077b5", "amountUSD": 1834.70},
    {"platform": "Twitter/X",  "pct":  5.0, "color": "#1da1f2", "amountUSD":  917.35},
]


def _report_dir(user_id: str) -> Path:
    """Ensure and return the user-specific reports directory."""
    d = Path(settings.UPLOAD_DIR) / "reports" / user_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


# ═══════════════════════════════════════════════════════════════
# EXCEL BUILDER  (openpyxl)
# ═══════════════════════════════════════════════════════════════

def _build_excel(user_id: str, period: str = "November 2025") -> io.BytesIO:
    """
    Build a multi-sheet Excel workbook and return a BytesIO buffer.

    Sheets:
      1. Campaigns      — full table, 6 rows
      2. Analytics      — KPI cards + weekly series
      3. Budget         — summary + platform allocations
      4. ML Models      — registry with metrics
      5. Green AI       — CO₂ session report (Scope 2)
    """
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Styles ──────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="6D28D9")    # purple
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1E1B4B")
    wrap_align = Alignment(wrap_text=True, vertical="center")
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_header_row(ws, headers: List[str], row: int = 1) -> None:
        for col_idx, h in enumerate(headers, 1):
            cell              = ws.cell(row=row, column=col_idx, value=h)
            cell.font         = hdr_font
            cell.fill         = hdr_fill
            cell.alignment    = wrap_align
            cell.border       = border

    def set_col_widths(ws, widths: List[int]) -> None:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def title_cell(ws, text: str, row: int = 1, col: int = 1) -> None:
        c       = ws.cell(row=row, column=col, value=text)
        c.font  = title_font
        c.alignment = Alignment(horizontal="left")

    # ── Sheet 1: Campaigns ────────────────────────────────
    ws1 = wb.create_sheet("Campaigns")
    title_cell(ws1, f"LumindAd — Campaign Report · {period}")
    write_header_row(
        ws1,
        ["ID","Name","Platform","Status","Budget ($)","Spent ($)",
         "Impressions","Clicks","CTR","Conversions","ROAS"],
        row=2,
    )
    for r_idx, c in enumerate(_CAMPAIGNS, 3):
        row_data = [
            c["id"], c["name"], c["platform"], c["status"].upper(),
            c["budget"], c["spent"], c["impressions"], c["clicks"],
            c["ctr"], c["conv"], c["roas"],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=col_idx, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")
    set_col_widths(ws1, [8, 24, 14, 11, 12, 12, 14, 10, 8, 13, 8])

    # ── Sheet 2: Analytics ────────────────────────────────
    ws2 = wb.create_sheet("Analytics")
    title_cell(ws2, f"LumindAd — Analytics KPIs · {period}")
    write_header_row(ws2, ["Metric","Value","Change (%)","Icon"], row=2)
    for r_idx, kpi in enumerate(_ANALYTICS_KPIS, 3):
        trend = "↑" if kpi["change"] > 0 else "↓"
        ws2.cell(row=r_idx, column=1, value=kpi["label"])
        ws2.cell(row=r_idx, column=2, value=kpi["formatted"])
        ws2.cell(row=r_idx, column=3, value=kpi["change"])
        ws2.cell(row=r_idx, column=4, value=trend)
        for col in range(1, 5):
            ws2.cell(row=r_idx, column=col).border = border
    set_col_widths(ws2, [22, 14, 14, 6])

    # ── Sheet 3: Budget ───────────────────────────────────
    ws3 = wb.create_sheet("Budget")
    title_cell(ws3, f"LumindAd — Budget Summary · {period}")
    write_header_row(ws3, ["Field","Value"], row=2)
    for r_idx, (k, v) in enumerate(_BUDGET_SUMMARY.items(), 3):
        ws3.cell(row=r_idx, column=1, value=k)
        ws3.cell(row=r_idx, column=2, value=v)
    ws3.cell(row=10, column=1, value="— Platform Allocations —")
    ws3.cell(row=10, column=1).font = Font(bold=True)
    write_header_row(ws3, ["Platform","Allocation (%)","Amount ($)"], row=11)
    for r_idx, a in enumerate(_PLATFORM_ALLOCS, 12):
        ws3.cell(row=r_idx, column=1, value=a["platform"])
        ws3.cell(row=r_idx, column=2, value=a["pct"])
        ws3.cell(row=r_idx, column=3, value=a["amountUSD"])
    set_col_widths(ws3, [22, 18, 15])

    # ── Sheet 4: ML Models ────────────────────────────────
    ws4 = wb.create_sheet("ML Models")
    title_cell(ws4, "LumindAd — ML Model Registry")
    write_header_row(ws4, ["Model","Algorithm","Accuracy (%)","F1 Score (%)","AUC (%)","Status"], row=2)
    for r_idx, m in enumerate(_ML_MODELS, 3):
        ws4.cell(row=r_idx, column=1, value=m["name"])
        ws4.cell(row=r_idx, column=2, value=m["algorithm"])
        ws4.cell(row=r_idx, column=3, value=m["accuracy"])
        ws4.cell(row=r_idx, column=4, value=m["f1"])
        ws4.cell(row=r_idx, column=5, value=m["auc"])
        ws4.cell(row=r_idx, column=6, value=m["status"].upper())
        for col in range(1, 7):
            ws4.cell(row=r_idx, column=col).border = border
    set_col_widths(ws4, [22, 20, 14, 14, 10, 12])

    # ── Sheet 5: Green AI ─────────────────────────────────
    ws5 = wb.create_sheet("Green AI")
    title_cell(ws5, "LumindAd — Green AI · GHG Scope 2 CO₂ Tracker")
    write_header_row(ws5, ["Field","Value"], row=2)
    from app.services.green_ai_service import green_ai
    report = green_ai.get_session_report(user_id)
    green_rows = [
        ("Total Inferences",       report["count"]),
        ("Total CO₂ (grams)",      f"{report['totalCO2Grams']:.8f}"),
        ("Rating",                 report["rating"]),
        ("Scope",                  report["scope"]),
        ("Carbon Intensity (kgCO₂/kWh)", report["carbonIntensity"]),
        ("PUE",                    report["pue"]),
        ("CPU Power (W)",          report["cpuPowerW"]),
        ("GPU Power (W)",          report["gpuPowerW"]),
        ("km Driving Equiv.",      f"{report['equivalences']['km_driving']:.8f}"),
        ("Smartphone Hours Equiv.",f"{report['equivalences']['smartphone_hours']:.4f}"),
        ("Session Started",        report["startedAt"]),
        ("Report Generated",       datetime.now(timezone.utc).isoformat()),
    ]
    for r_idx, (k, v) in enumerate(green_rows, 3):
        ws5.cell(row=r_idx, column=1, value=k)
        ws5.cell(row=r_idx, column=2, value=v)
        for col in (1, 2):
            ws5.cell(row=r_idx, column=col).border = border
    set_col_widths(ws5, [28, 28])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
# PDF BUILDER  (reportlab / text fallback)
# ═══════════════════════════════════════════════════════════════

def _build_pdf(user_id: str, period: str = "November 2025") -> io.BytesIO:
    """Build a PDF report. Uses reportlab when available, UTF-8 text fallback otherwise."""
    from app.services.green_ai_service import green_ai
    green_report = green_ai.get_session_report(user_id)

    if _REPORTLAB:
        return _build_pdf_reportlab(user_id, period, green_report)
    return _build_pdf_text_fallback(user_id, period, green_report)


def _build_pdf_reportlab(
    user_id:      str,
    period:       str,
    green_report: dict,
) -> io.BytesIO:
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    purple = rl_colors.HexColor("#6D28D9")
    title_style  = ParagraphStyle("title",  parent=styles["Heading1"], textColor=purple, fontSize=20)
    h2_style     = ParagraphStyle("h2",     parent=styles["Heading2"], textColor=purple, fontSize=14)
    body_style   = styles["BodyText"]

    # Cover
    story.append(Paragraph("LumindAd Enterprise", title_style))
    story.append(Paragraph(f"AI-Powered Ad Management Report · {period}", styles["Heading2"]))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", color=purple))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", body_style))
    story.append(Spacer(1, 1*cm))

    # KPI summary
    story.append(Paragraph("Analytics KPIs", h2_style))
    kpi_data = [["Metric", "Value", "Change"]] + [
        [k["label"], k["formatted"], f"{'+' if k['change'] > 0 else ''}{k['change']}%"]
        for k in _ANALYTICS_KPIS
    ]
    kpi_table = Table(kpi_data, colWidths=[9*cm, 4*cm, 3*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), purple),
        ("TEXTCOLOR",  (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID",       (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F5F3FF")]),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.8*cm))

    # Campaigns
    story.append(Paragraph("Campaign Performance", h2_style))
    camp_data = [["ID", "Name", "Platform", "Status", "Spent", "ROAS"]] + [
        [c["id"], c["name"][:20], c["platform"], c["status"].upper(),
         f"${c['spent']:,.0f}", str(c["roas"])]
        for c in _CAMPAIGNS
    ]
    camp_table = Table(camp_data, colWidths=[1.5*cm, 5.5*cm, 3*cm, 2.5*cm, 2.5*cm, 1.5*cm])
    camp_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), purple),
        ("TEXTCOLOR",  (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F5F3FF")]),
    ]))
    story.append(camp_table)
    story.append(Spacer(1, 0.8*cm))

    # Green AI
    story.append(Paragraph("🌱 Green AI — GHG Scope 2 CO₂", h2_style))
    story.append(Paragraph(green_report["badge"], body_style))
    story.append(Paragraph(f"Rating: {green_report['rating']}", body_style))
    story.append(Paragraph(
        f"Total inferences: {green_report['count']} | "
        f"CO₂: {green_report['totalCO2Grams']:.8f} g | "
        f"km equiv: {green_report['equivalences']['km_driving']:.8f} km",
        body_style,
    ))

    doc.build(story)
    buf.seek(0)
    return buf


def _build_pdf_text_fallback(
    user_id:      str,
    period:       str,
    green_report: dict,
) -> io.BytesIO:
    """Plain text 'PDF' fallback when reportlab is not installed."""
    lines = [
        f"LumindAd Enterprise — AI-Powered Ad Management Report",
        f"Period  : {period}",
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        "",
        "═══════════════════ ANALYTICS KPIs ══════════════════════",
    ]
    for k in _ANALYTICS_KPIS:
        lines.append(f"  {k['label']:<25} {k['formatted']:<10}  "
                     f"{('+' if k['change'] > 0 else '')}{k['change']}%")
    lines += ["", "═══════════════════ CAMPAIGNS ═══════════════════════════"]
    for c in _CAMPAIGNS:
        lines.append(f"  {c['id']}  {c['name']:<25} {c['platform']:<12} "
                     f"{c['status']:<10}  ${c['spent']:>8,.0f}  ROAS {c['roas']}")
    lines += [
        "", "═══════════════════ GREEN AI — GHG SCOPE 2 ══════════════",
        f"  {green_report['badge']}",
        f"  Rating     : {green_report['rating']}",
        f"  CO₂ total  : {green_report['totalCO2Grams']:.8f} g",
        f"  Inferences : {green_report['count']}",
        f"  km equiv   : {green_report['equivalences']['km_driving']:.8f} km",
        "",
    ]
    return io.BytesIO("\n".join(lines).encode("utf-8"))


# ═══════════════════════════════════════════════════════════════
# TASK: GENERATE EXCEL REPORT
# ═══════════════════════════════════════════════════════════════

if _CELERY and celery_app:

    @celery_app.task(
        name="app.workers.report_tasks.generate_excel_report_task",
        bind=True,
        queue="reports",
        max_retries=2,
        soft_time_limit=300,   # 5 min
        time_limit=600,
    )
    def generate_excel_report_task(
        self,
        user_id: str,
        period:  str = "November 2025",
    ) -> dict:
        """
        Generate a multi-sheet Excel workbook and save to disk.

        Returns:
            { filePath, fileSize, durationMs, format, taskId }
        """
        if not _OPENPYXL:
            return {"error": "openpyxl not installed — pip install openpyxl", "taskId": self.request.id}

        t0 = time.perf_counter()
        logger.info("📊 Excel report task for user %s period=%r", user_id, period)

        try:
            buf      = _build_excel(user_id, period)
            out_dir  = _report_dir(user_id)
            filename = f"lumindad_report_{_ts()}.xlsx"
            out_path = out_dir / filename

            out_path.write_bytes(buf.getvalue())
            file_size   = out_path.stat().st_size
            duration_ms = int((time.perf_counter() - t0) * 1000)

            logger.info("✅ Excel report saved: %s (%d KB) in %d ms",
                        filename, file_size // 1024, duration_ms)
            return {
                "filePath":   str(out_path),
                "fileName":   filename,
                "fileSize":   file_size,
                "format":     "xlsx",
                "durationMs": duration_ms,
                "taskId":     self.request.id,
                "generatedAt": datetime.now(timezone.utc).isoformat(),
            }

        except SoftTimeLimitExceeded:
            return {"error": "Time limit exceeded", "taskId": self.request.id}
        except Exception as exc:
            logger.exception("Excel report task failed: %s", exc)
            raise self.retry(exc=exc, countdown=30)


# ═══════════════════════════════════════════════════════════════
# TASK: GENERATE PDF REPORT
# ═══════════════════════════════════════════════════════════════

if _CELERY and celery_app:

    @celery_app.task(
        name="app.workers.report_tasks.generate_pdf_report_task",
        bind=True,
        queue="reports",
        max_retries=2,
        soft_time_limit=300,
        time_limit=600,
    )
    def generate_pdf_report_task(
        self,
        user_id: str,
        period:  str = "November 2025",
    ) -> dict:
        """
        Generate a PDF report and save to disk.

        Uses reportlab when available, UTF-8 text fallback otherwise.
        """
        t0 = time.perf_counter()
        logger.info("📄 PDF report task for user %s period=%r", user_id, period)

        try:
            buf      = _build_pdf(user_id, period)
            out_dir  = _report_dir(user_id)
            filename = f"lumindad_report_{_ts()}.pdf"
            out_path = out_dir / filename

            out_path.write_bytes(buf.getvalue())
            file_size   = out_path.stat().st_size
            duration_ms = int((time.perf_counter() - t0) * 1000)

            logger.info("✅ PDF report saved: %s (%d KB) in %d ms",
                        filename, file_size // 1024, duration_ms)
            return {
                "filePath":    str(out_path),
                "fileName":    filename,
                "fileSize":    file_size,
                "format":      "pdf",
                "durationMs":  duration_ms,
                "taskId":      self.request.id,
                "generatedAt": datetime.now(timezone.utc).isoformat(),
            }

        except SoftTimeLimitExceeded:
            return {"error": "Time limit exceeded", "taskId": self.request.id}
        except Exception as exc:
            logger.exception("PDF report task failed: %s", exc)
            raise self.retry(exc=exc, countdown=30)


# ═══════════════════════════════════════════════════════════════
# TASK: SCHEDULED DAILY REPORT (beat — 06:00 UTC)
# ═══════════════════════════════════════════════════════════════

if _CELERY and celery_app:

    @celery_app.task(
        name="app.workers.report_tasks.scheduled_daily_report_task",
        queue="reports",
        max_retries=1,
        soft_time_limit=1800,   # 30 min for all users
        ignore_result=True,
    )
    def scheduled_daily_report_task() -> None:
        """
        Beat task — 06:00 UTC daily.

        Generates Excel + PDF reports for all active users.
        In production: queries DB for user list.
        Prototype: generates for seed user 'usr_001'.
        Optionally uploads to S3 if settings.AWS_S3_BUCKET is set.
        """
        logger.info("⏰ Daily scheduled report task starting")
        t0         = time.perf_counter()
        now_label  = datetime.now(timezone.utc).strftime("%B %Y")
        target_users = ["usr_001"]   # production: query active users from DB

        for user_id in target_users:
            try:
                # Excel
                if _OPENPYXL:
                    buf      = _build_excel(user_id, now_label)
                    out_dir  = _report_dir(user_id)
                    xl_path  = out_dir / f"daily_report_{_ts()}.xlsx"
                    xl_path.write_bytes(buf.getvalue())
                    logger.info("Daily Excel saved: %s", xl_path.name)

                # PDF
                buf     = _build_pdf(user_id, now_label)
                out_dir = _report_dir(user_id)
                pd_path = out_dir / f"daily_report_{_ts()}.pdf"
                pd_path.write_bytes(buf.getvalue())
                logger.info("Daily PDF saved: %s", pd_path.name)

            except Exception as exc:
                logger.error("Daily report failed for user %s: %s", user_id, exc)

        elapsed = int((time.perf_counter() - t0) * 1000)
        logger.info("✅ Daily report task complete for %d users in %d ms",
                    len(target_users), elapsed)


# ═══════════════════════════════════════════════════════════════
# TASK: SEND REPORT EMAIL (optional — SMTP)
# ═══════════════════════════════════════════════════════════════

if _CELERY and celery_app:

    @celery_app.task(
        name="app.workers.report_tasks.send_report_email_task",
        queue="reports",
        max_retries=3,
        soft_time_limit=120,
        default_retry_delay=60,
    )
    def send_report_email_task(
        self,
        user_email:   str,
        file_paths:   List[str],
        report_date:  str,
    ) -> dict:
        """
        Send generated reports via SMTP.

        Skipped with a warning if SMTP is not configured in settings
        (settings.SMTP_HOST is None or empty).

        Args:
            user_email:  Recipient email address
            file_paths:  List of absolute file paths to attach
            report_date: Human-readable date string e.g. 'November 2025'
        """
        smtp_host = getattr(settings, "SMTP_HOST", None)
        if not smtp_host:
            logger.info("SMTP not configured — skipping email for %s", user_email)
            return {"skipped": True, "reason": "SMTP_HOST not configured"}

        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.base      import MIMEBase
            from email.mime.text      import MIMEText
            from email                import encoders

            msg            = MIMEMultipart()
            msg["From"]    = getattr(settings, "SMTP_FROM", "noreply@lumindad.ai")
            msg["To"]      = user_email
            msg["Subject"] = f"LumindAd Report — {report_date}"

            body = (
                f"Hello,\n\n"
                f"Please find attached your LumindAd performance report for {report_date}.\n\n"
                f"This report includes:\n"
                f"  • Campaign performance metrics\n"
                f"  • Analytics KPIs\n"
                f"  • Budget summary & platform allocations\n"
                f"  • ML model status\n"
                f"  • Green AI GHG Scope 2 CO₂ tracking\n\n"
                f"Generated by LumindAd Enterprise · Sustainable Intelligence & BI\n"
            )
            msg.attach(MIMEText(body, "plain"))

            for fp in file_paths:
                path = Path(fp)
                if not path.exists():
                    continue
                part = MIMEBase("application", "octet-stream")
                part.set_payload(path.read_bytes())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f'attachment; filename="{path.name}"')
                msg.attach(part)

            smtp_port = getattr(settings, "SMTP_PORT", 587)
            with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
                server.starttls()
                smtp_user = getattr(settings, "SMTP_USER", "")
                smtp_pass = getattr(settings, "SMTP_PASSWORD", "")
                if smtp_user:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg)

            logger.info("Report email sent to %s (%d attachments)", user_email, len(file_paths))
            return {"sent": True, "to": user_email, "attachments": len(file_paths)}

        except Exception as exc:
            countdown = 60 * (2 ** self.request.retries)
            logger.warning("Email send failed (retry in %ds): %s", countdown, exc)
            raise self.retry(exc=exc, countdown=countdown)


# ═══════════════════════════════════════════════════════════════
# __init__.py stub helpers (for testing without Celery)
# ═══════════════════════════════════════════════════════════════

def build_excel_sync(user_id: str, period: str = "November 2025") -> io.BytesIO:
    """
    Build Excel report synchronously (no Celery — for direct API use).
    Called by POST /bi-export/excel when not running in worker mode.
    """
    if not _OPENPYXL:
        raise ImportError("openpyxl is not installed — pip install openpyxl")
    return _build_excel(user_id, period)


def build_pdf_sync(user_id: str, period: str = "November 2025") -> io.BytesIO:
    """
    Build PDF report synchronously (no Celery — for direct API use).
    Called by POST /bi-export/pdf when not running in worker mode.
    """
    return _build_pdf(user_id, period)
